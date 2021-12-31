import sys

import openpyxl as xl
import shutil
import oracle
from datetime import datetime


SETTINGS = {
            "start":    {"A+": 4, "A-": 10, "Q1": 16, "Q2": 22, "Q3": 28, "Q4": 34},
            "end":    {"A+": 5, "A-": 11, "Q1": 17, "Q2": 23, "Q3": 29, "Q4": 35},
            }


# Функции для получения дат начала и конца периода
def check_dates_difference(first_date, second_date):
    """
    Проверяет не равные ли даты
    Args:
        first_date, second_date: Даты введенные пользователями
    Returns:
        True - если они равны, если это не так False
    """
    if first_date == second_date:
        print('Две даты отчетного периода должны быть разными.')
        return True
    else:
        return False


def request_date(txt: str) -> datetime:
    """
    Запрашиваем у пользователя дату в формате ДД-ММ-ГГГГ
    Args:
        txt (str): Текст выдаваемый запросом
             Правильность введеной даты проверяется фунцией
             get_date(str_date)
    Returns:
        date (datetime): Преобразованая дата
    """
    true_date = False
    date = None
    while not true_date:
        print(txt)
        str_date = input('В формате ДД-ММ-ГГГГ или q для выхода:')
        if str_date == 'q':
            sys.exit()
        true_date, date = check_validation_date(str_date)

    return date


class FutureDate(Exception):
    """
    Обработчик исключения. Дата конца и начала периода,
    должны быть раньше сегодня
    """

    def __init__(self, text):
        self.txt = text


def check_validation_date(str_date: str) -> tuple:
    """
    Переводим str_date в дату, попутно проверяем ее:
        - Что бы дата была раньше сегодня
        - Что бы была в правильном формате
        - Что бы существовала
    Args:
        str_date (str): дата в виде строки, должна быть в формате "ДД-ММ-ГГГГ"
    Returns:
        True или False - в зависимости от того, смогли ли мы преобразовать дату
        date - преобразованую строку в datetime
    """
    date_valid = False
    date = datetime.strptime('01-01-1900', '%d-%m-%Y')
    try:
        date = datetime.strptime(str_date, '%d-%m-%Y')
        if date > datetime.now():
            raise FutureDate('Даты должны быть в прошлом')
        date_valid = True

    except ValueError:
        print('Вы ввели дату или не в формате ДД-ММ-ГГГГ.')
        print('Или не существующую дату')
    except FutureDate as mr:
        print(mr)

    finally:
        print()
        return date_valid, date


def order_dates(first_date, second_date):
    """
    Сортируем даты, и возращаем их
    Args:
        first_date, second_date:  Даты введенные пользователями
    Returns:
        Возращает отсортированые даты
    """
    return sorted([first_date, second_date])


# File from template
def new_file_name(template, since_date, to_date):
    """
    Create name for new file = region name + month
        region name - we get from template name
        month - number of month
    :param template: Source file - file with template
    :param to_date: date of end  of period
    :return:
        {region_name}_{month:02}_{year:04}.xlsx
    """
    template_name = template.split('_')[-1]
    first_date = f'{since_date.day:02}.{since_date.month:02}.{since_date.year:04}'
    second_date = f'{to_date.day:02}.{to_date.month:02}.{to_date.year:04}'
    return f'./{template_name}_{first_date}-{second_date}.xlsx'


def copy_template_file(template_file, new_file):
    """
    Копируем шаблон <template_file> в новый файл <new_file>
    Args:
        template_file (str): Имя одно из файлов шаблона, которые находятся в папке Templates
        new_file (str): Имя нового файла сформированого из шаблон
    Returns:
        True - Если копия создана, False если произошла ошибка
    """
    try:
        shutil.copy(template_file, new_file)
        return True

    except shutil.SameFileError:
        # Исключение на будущее, когда можно будет задавать имя нового файла через аргументы
        print('Имя Шаблона и нового файла совпадают')
    except PermissionError:
        # Исключение при недоступноти
        print("Проблемы с доступом. Обратитесть к Администратору")
    except Exception:
        # For other errors
        print("Error occurred while copying file.")
    return False


def get_pids(template_file, pid_column=1):
    """
    Получаем points_id с ексел файла. Проверяя колонку под номером <pid_column>
    :param template_file: файл шаблона
    :param pid_column: Колонка с points_id. Default value = 1
    :return pids: diction {point id: номер колонки}
    """
    wb, ws = open_xlsx(template_file)

    # calculate total number of rows and
    # columns in source excel file
    mr, mc = ws.max_row, ws.max_column

    pids = dict()
    for num_row in range(1, mr + 1):
        pid = ws.cell(row=num_row, column=pid_column).value
        if pid:
            pids[pid] = num_row
    wb.close()

    return pids


def open_xlsx(file_name):
    wb = xl.load_workbook(file_name)
    ws = wb.worksheets[0]
    return wb, ws


# Put data from DB into xlsx file
def compare_cell_2_data(pid_row, raw_data_since, raw_data_to):
    dict_row_data_since = dict()
    dict_row_data_to = dict()
    for data_line in raw_data_since:
        row = pid_row.get(data_line[0])
        if row:
            dict_row_data_since[row] = {"A+": data_line[2], "A-": data_line[3],
                                        "Q1": data_line[4], "Q3": data_line[5]}
    for data_line in raw_data_to:
        row = pid_row.get(data_line[0])
        if row:
            dict_row_data_to[row] = {"A+": data_line[2], "A-": data_line[3],
                                     "Q1": data_line[4], "Q3": data_line[5]}
    return dict_row_data_since, dict_row_data_to


def create_file_with_data(file_name: str, dict_row_data_since, dict_row_data_to, column = 1):

    wb, ws = open_xlsx(file_name)

    for row in range(1, ws.max_row + 1):
        ws.cell(row=row, column=column).value = None

        if dict_row_data_since.get(row):
            fill_xls_row(ws, row, SETTINGS["start"], dict_row_data_since[row])

        if dict_row_data_to.get(row):
            fill_xls_row(ws, row, SETTINGS["end"], dict_row_data_to[row])


    wb.save(file_name)
    wb.close()


def fill_xls_row(worksheet, row, value_position, dict_row_data):

    for param, param_value in dict_row_data.items():
        worksheet.cell(row=row, column=value_position[param]).value = param_value



